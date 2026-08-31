"""Normalize HUD FY2026 SAFMR bulk xlsx → ZIP×bedroom observations + ZCTA layer."""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from shapely.geometry import mapping, shape
from shapely.strtree import STRtree

from rent_seekers.compare.calculate import build_comparison
from rent_seekers.config import load_yaml, project_root
from rent_seekers.geography.simplify import process_feature_collection
from rent_seekers.models import MarketRentObservation, MeasureBasis, TenantRentObservation
from rent_seekers.sources import hud_safmr as safmr_source
from rent_seekers.sources import zcta as zcta_source
from rent_seekers.sources.base import sha256_file, utc_now, write_json

BEDROOM_KEYS = (0, 1, 2, 3, 4)
BR_LABEL = {0: "0br", 1: "1br", 2: "2br", 3: "3br", 4: "4br"}


def policy() -> dict[str, Any]:
    return load_yaml(project_root() / "config" / "hud_safmr.yml")


def _processed_root() -> Path:
    return project_root() / "data" / "processed" / "hud_safmr"


def _public_mirrors() -> list[Path]:
    root = project_root()
    return [
        _processed_root(),
        root / "web" / "public" / "data" / "hud_safmr",
        root / "dist" / "data" / "hud_safmr",
        root / "dist" / "app" / "data" / "hud_safmr",
    ]


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def _zip5(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # openpyxl may yield int
    if s.isdigit():
        return s.zfill(5)
    m = re.search(r"(\d{5})", s)
    return m.group(1) if m else None


def _header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, col in enumerate(header_row):
        key = _norm_header(col)
        if key:
            out[key] = i
    return out


def _resolve_col(hmap: dict[str, int], *candidates: str) -> int:
    for c in candidates:
        n = _norm_header(c)
        if n in hmap:
            return hmap[n]
        # try looser match
        for k, idx in hmap.items():
            if n.replace(" ", "") == k.replace(" ", ""):
                return idx
    raise KeyError(f"column not found among {candidates}; have {list(hmap)[:12]}")


def parse_safmr_xlsx(
    path: Path,
    *,
    cfg: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """
    Parse bulk SAFMR xlsx into per-ZIP bedroom rent rows.

    Filters to the New York HUD Metro FMR Area when configured.
    """
    cfg = cfg or policy()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter)
        hmap = _header_map(tuple(header))
        zip_i = _resolve_col(hmap, cfg.get("zip_column") or "ZIP Code", "ZIP Code")
        area_name_i = _resolve_col(
            hmap,
            cfg.get("area_name_column") or "HUD Fair Market Rent Area Name",
            "HUD Fair Market Rent Area Name",
        )
        area_code_i = _resolve_col(
            hmap,
            cfg.get("area_code_column") or "HUD Area Code",
            "HUD Area Code",
        )
        br_cols: dict[int, int] = {}
        br_cfg = cfg.get("bedroom_columns") or {}
        for br in BEDROOM_KEYS:
            # config may use int keys (yaml) or string keys
            cand = br_cfg.get(br) or br_cfg.get(str(br)) or f"SAFMR {br}BR"
            # Prefer short labels after normalization: "SAFMR 0BR"
            short = f"SAFMR {br}BR"
            try:
                br_cols[br] = _resolve_col(hmap, str(cand), short)
            except KeyError:
                # fall back: find key starting with SAFMR NBR
                found = None
                for k, idx in hmap.items():
                    if k.startswith(short) or k.replace(" ", "").startswith(
                        short.replace(" ", "")
                    ):
                        found = idx
                        break
                if found is None:
                    raise
                br_cols[br] = found

        area_filter = (cfg.get("hud_area_name_contains") or "").strip()
        area_code_filter = (cfg.get("hud_area_code") or "").strip()
        out: list[dict[str, Any]] = []
        for raw in rows_iter:
            if not raw:
                continue
            area_name = str(raw[area_name_i] or "").strip()
            area_code = str(raw[area_code_i] or "").strip()
            if area_filter and area_filter not in area_name:
                continue
            if area_code_filter and area_code != area_code_filter and area_filter:
                # name filter already applied; allow code mismatch only if name matched
                pass
            zip_code = _zip5(raw[zip_i])
            if not zip_code:
                continue
            bedrooms: dict[str, int] = {}
            for br, col_i in br_cols.items():
                val = raw[col_i]
                if val is None or val == "":
                    continue
                try:
                    rent = int(round(float(val)))
                except (TypeError, ValueError):
                    continue
                if rent <= 0:
                    continue
                bedrooms[str(br)] = rent
            if not bedrooms:
                continue
            out.append(
                {
                    "zip": zip_code,
                    "hud_area_code": area_code,
                    "hud_area_name": area_name,
                    "bedrooms": bedrooms,
                }
            )
        return out
    finally:
        wb.close()


def _observation_id(zip_code: str, bedroom: int, fiscal_year: str = "FY2026") -> str:
    fy = fiscal_year.lower().replace(" ", "")
    return f"hud-safmr:{fy}:{zip_code}:{BR_LABEL[bedroom]}"


def build_market_observations(
    zip_rows: list[dict[str, Any]],
    *,
    cfg: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """Build MarketRentObservation dicts for every ZIP × bedroom."""
    cfg = cfg or policy()
    measure = cfg.get("measure") or {}
    period_start = date.fromisoformat(cfg["period_start"])
    period_end = date.fromisoformat(cfg["period_end"])
    fiscal_year = cfg.get("fiscal_year") or "FY2026"
    artifact_id = cfg.get("artifact_id") or safmr_source.ARTIFACT_ID
    source_url = safmr_source.source_cfg().get("landing_page")
    label = cfg.get("display_label") or (
        "HUD FY2026 Small Area Fair Market Rent — ZIP-level gross-rent benchmark"
    )
    not_a = cfg.get("not_a_label") or "median asking rent"

    obs: list[dict[str, Any]] = []
    for row in zip_rows:
        z = row["zip"]
        for br_s, value in (row.get("bedrooms") or {}).items():
            br = int(br_s)
            obs.append(
                MarketRentObservation(
                    observation_id=_observation_id(z, br, fiscal_year),
                    market_area_id=f"zcta:{z}",
                    period_start=period_start,
                    period_end=period_end,
                    measure_basis=MeasureBasis.regulatory_market_benchmark,
                    gross_or_net=str(measure.get("gross_or_net") or "gross"),
                    statistic=str(
                        measure.get("statistic") or "40th_percentile_methodology"
                    ),
                    unit_scope=str(measure.get("unit_scope") or "bedroom_specific"),
                    bedroom_count=br,
                    currency="USD",
                    cadence="monthly",
                    value=float(value),
                    sample_size=None,
                    source_artifact_id=artifact_id,
                    source_url=source_url,
                    notes=(
                        f"{label}. Not {not_a}. "
                        f"ZIP/ZCTA {z}; bedroom={br}; fiscal_year={fiscal_year}; "
                        f"hud_area={row.get('hud_area_name')}"
                    ),
                ).model_dump(mode="json")
            )
    return obs


def build_market_areas(zip_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "market_area_id": f"zcta:{row['zip']}",
            "geography_type": "zcta",
            "name": row["zip"],
            "vintage": "2020",
            "geometry_id": f"zcta:{row['zip']}:2020",
        }
        for row in zip_rows
    ]


def build_zcta_choropleth(
    *,
    zip_rows: list[dict[str, Any]],
    zcta_fc: dict[str, Any] | None = None,
    cfg: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Join SAFMR values onto ZCTA polygons. Missing ZIPs stay visible with
    `safmr_missing=true` so the layer can render gaps.
    """
    cfg = cfg or policy()
    if zcta_fc is None:
        zcta_fc = zcta_source.load_raw()
    geo_cfg = load_yaml(project_root() / "config" / "geography.yml")
    tol = float((geo_cfg.get("simplify") or {}).get("zctas") or 0.00018)

    by_zip = {r["zip"]: r for r in zip_rows}
    fiscal_year = cfg.get("fiscal_year") or "FY2026"
    period_start = cfg.get("period_start")
    period_end = cfg.get("period_end")
    artifact_id = cfg.get("artifact_id") or safmr_source.ARTIFACT_ID
    source_url = safmr_source.source_cfg().get("landing_page")
    label = cfg.get("display_label") or (
        "HUD FY2026 Small Area Fair Market Rent — ZIP-level gross-rent benchmark"
    )
    not_a = cfg.get("not_a_label") or "median asking rent"

    matched = 0
    missing = 0

    def prop_xform(props: dict[str, Any], geom: Any, index: int) -> dict[str, Any] | None:
        nonlocal matched, missing
        del geom, index
        z = _zip5(props.get("zcta5") or props.get("ZCTA5") or props.get("zip"))
        if not z:
            return None
        row = by_zip.get(z)
        bedrooms = (row or {}).get("bedrooms") or {}
        has = bool(bedrooms)
        if has:
            matched += 1
        else:
            missing += 1
        out_props: dict[str, Any] = {
            "market_area_id": f"zcta:{z}",
            "zcta": z,
            "zip": z,
            "geography_type": "zcta",
            "name": z,
            "vintage": "2020",
            "fiscal_year": fiscal_year,
            "period_start": period_start,
            "period_end": period_end,
            "measure_basis": "regulatory_market_benchmark",
            "gross_or_net": "gross",
            "statistic": "40th_percentile_methodology",
            "source_id": safmr_source.SOURCE_ID,
            "source_artifact_id": artifact_id,
            "source_url": source_url,
            "source_label": label,
            "not_a_label": not_a,
            "safmr_missing": not has,
            "safmr_0br": bedrooms.get("0"),
            "safmr_1br": bedrooms.get("1"),
            "safmr_2br": bedrooms.get("2"),
            "safmr_3br": bedrooms.get("3"),
            "safmr_4br": bedrooms.get("4"),
            # Default display field for choropleth (UI swaps bedroom via expression)
            "safmr_rent_usd": bedrooms.get("2"),
            "bedroom_default": 2,
        }
        if row:
            out_props["hud_area_code"] = row.get("hud_area_code")
            out_props["hud_area_name"] = row.get("hud_area_name")
        return out_props

    polygons, _points, _review = process_feature_collection(
        zcta_fc,
        tolerance=tol,
        property_transform=prop_xform,
        include_points=False,
    )

    return {
        "type": "FeatureCollection",
        "features": polygons.get("features") or [],
        "meta": {
            "matched_zctas": matched,
            "missing_zctas": missing,
            "feature_count": len(polygons.get("features") or []),
            "fiscal_year": fiscal_year,
            "period_start": period_start,
            "period_end": period_end,
            "source_id": safmr_source.SOURCE_ID,
            "source_artifact_id": artifact_id,
            "display_label": label,
            "not_a_label": not_a,
            "gross_or_net": "gross",
            "bedroom_keys": list(BEDROOM_KEYS),
        },
    }


def assign_developments_to_zcta(
    developments_fc: dict[str, Any],
    zcta_fc: dict[str, Any],
) -> list[dict[str, Any]]:
    """
    Polygon → ZCTA via representative-point containment, then largest overlap.

    Returns geography_assignment records (§6.6 shape).
    """
    zcta_geoms: list[Any] = []
    zcta_zips: list[str] = []
    for feat in zcta_fc.get("features") or []:
        props = feat.get("properties") or {}
        z = _zip5(props.get("zcta") or props.get("zcta5") or props.get("zip"))
        geom = feat.get("geometry")
        if not z or not geom:
            continue
        try:
            g = shape(geom)
        except Exception:
            continue
        if g.is_empty:
            continue
        if not g.is_valid:
            g = g.buffer(0)
        zcta_geoms.append(g)
        zcta_zips.append(z)

    if not zcta_geoms:
        return []

    tree = STRtree(zcta_geoms)
    # shapely 2: tree.query returns indices
    assignments: list[dict[str, Any]] = []

    for feat in developments_fc.get("features") or []:
        props = feat.get("properties") or {}
        did = props.get("development_id")
        geom = feat.get("geometry")
        if not did or not geom:
            continue
        try:
            dev_g = shape(geom)
        except Exception:
            continue
        if dev_g.is_empty:
            continue
        if not dev_g.is_valid:
            dev_g = dev_g.buffer(0)
        try:
            pt = dev_g.representative_point()
        except Exception:
            continue

        # Candidate ZCTAs intersecting the development (Shapely 2 returns ndarray indices)
        cand_indices: list[int] = []
        try:
            idxs = tree.query(dev_g, predicate="intersects")
            cand_indices = [int(i) for i in list(idxs)]
        except Exception:
            try:
                idxs = tree.query(dev_g)
                cand_indices = [int(i) for i in list(idxs)]
            except Exception:
                cand_indices = []
        if not cand_indices:
            try:
                idxs2 = tree.query(pt)
                cand_indices = [int(i) for i in list(idxs2)]
            except Exception:
                cand_indices = []

        best_zip: str | None = None
        best_share = 0.0
        method = "none"
        # Prefer containment of representative point
        for i in cand_indices:
            if i < 0 or i >= len(zcta_geoms):
                continue
            zg = zcta_geoms[i]
            z = zcta_zips[i]
            try:
                if zg.contains(pt) or zg.covers(pt):
                    best_zip = z
                    best_share = 1.0
                    method = "representative_point_in_zcta"
                    break
            except Exception:
                continue
        if best_zip is None:
            dev_area = float(dev_g.area) or 0.0
            for i in cand_indices:
                if i < 0 or i >= len(zcta_geoms):
                    continue
                zg = zcta_geoms[i]
                z = zcta_zips[i]
                try:
                    inter = dev_g.intersection(zg)
                    share = (float(inter.area) / dev_area) if dev_area > 0 else 0.0
                except Exception:
                    share = 0.0
                if share > best_share:
                    best_share = share
                    best_zip = z
                    method = "polygon_area_dominant"

        if best_zip is None:
            assignments.append(
                {
                    "assignment_id": f"{did}->zcta:unassigned",
                    "subject_id": did,
                    "geography_id": None,
                    "assignment_method": "unassigned",
                    "overlap_share": 0.0,
                    "is_primary": False,
                    "quality": "unavailable",
                    "source_geometry_vintages": {
                        "development": "nycha-open-data-current",
                        "geography": "2020",
                    },
                }
            )
            continue

        quality = (
            "strong"
            if best_share >= 0.5 or method.startswith("representative")
            else "representative"
        )
        assignments.append(
            {
                "assignment_id": f"{did}->zcta:{best_zip}",
                "subject_id": did,
                "geography_id": f"zcta:{best_zip}",
                "zcta": best_zip,
                "assignment_method": method,
                "overlap_share": round(float(best_share), 4),
                "is_primary": True,
                "quality": quality,
                "source_geometry_vintages": {
                    "development": "nycha-open-data-current",
                    "geography": "2020",
                },
            }
        )
    return assignments


def build_hud_comparisons(
    *,
    tenant_rents: list[dict[str, Any]],
    market_obs: list[dict[str, Any]],
    assignments: list[dict[str, Any]],
    bedroom: int = 2,
) -> list[dict[str, Any]]:
    """Build rent_comparison records via shared comparison code for one bedroom."""
    by_zip_br: dict[tuple[str, int], dict[str, Any]] = {}
    for m in market_obs:
        mid = m.get("market_area_id") or ""
        if not mid.startswith("zcta:"):
            continue
        z = mid.split(":", 1)[1]
        br = m.get("bedroom_count")
        if br is None:
            continue
        by_zip_br[(z, int(br))] = m

    assign_by_dev = {
        a["subject_id"]: a
        for a in assignments
        if a.get("subject_id") and a.get("zcta")
    }

    out: list[dict[str, Any]] = []
    for tr in tenant_rents:
        did = tr.get("housing_development_id")
        if not did:
            continue
        asn = assign_by_dev.get(did)
        if not asn:
            continue
        z = asn["zcta"]
        mraw = by_zip_br.get((z, bedroom))
        if not mraw:
            continue
        try:
            tenant = TenantRentObservation.model_validate(tr)
            market = MarketRentObservation.model_validate(mraw)
        except Exception:
            continue
        br_label = BR_LABEL[bedroom]
        comp = build_comparison(
            comparison_id=f"{did}__hud-safmr:fy2026:{z}:{br_label}",
            housing_development_id=str(did),
            tenant=tenant,
            market=market,
            extra_quality_reasons=[
                f"market geography is ZIP/ZCTA {z} (source-native HUD SAFMR), "
                "not the exact development footprint",
                "HUD SAFMR is a regulatory gross-rent benchmark, not median asking rent",
            ],
        )
        dump = comp.model_dump(mode="json")
        dump["market_source"] = "hud_safmr"
        dump["market_zcta"] = z
        dump["market_bedroom_count"] = bedroom
        dump["assignment_method"] = asn.get("assignment_method")
        out.append(dump)
    return out


def normalize(
    *,
    xlsx_path: Path | None = None,
    force_ingest: bool = False,
    write: bool = True,
    developments_fc: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Ingest (if needed) → parse → ZCTA join → write processed artifacts."""
    cfg = policy()
    if force_ingest or xlsx_path is None:
        receipt = safmr_source.ingest(force=force_ingest)
    else:
        receipt = {
            "artifact_id": cfg.get("artifact_id") or safmr_source.ARTIFACT_ID,
            "source_id": safmr_source.SOURCE_ID,
            "cache": "provided",
        }
    path = xlsx_path or safmr_source.load_raw_path()
    zcta_receipt = zcta_source.ingest(force=force_ingest)

    zip_rows = parse_safmr_xlsx(path, cfg=cfg)
    market_obs = build_market_observations(zip_rows, cfg=cfg)
    market_areas = build_market_areas(zip_rows)
    zcta_raw = zcta_source.load_raw()
    choropleth = build_zcta_choropleth(zip_rows=zip_rows, zcta_fc=zcta_raw, cfg=cfg)

    # Load developments for assignment when available
    if developments_fc is None:
        for rel in (
            "data/processed/geometry/developments.geojson",
            "web/public/data/geometry/developments.geojson",
        ):
            p = project_root() / rel
            if p.exists():
                with p.open(encoding="utf-8") as fh:
                    developments_fc = json.load(fh)
                break
    assignments: list[dict[str, Any]] = []
    if developments_fc:
        assignments = assign_developments_to_zcta(developments_fc, choropleth)

    # Compact by-zip lookup for the UI bedroom selector
    by_zip = {
        r["zip"]: {
            "bedrooms": r["bedrooms"],
            "hud_area_code": r.get("hud_area_code"),
            "hud_area_name": r.get("hud_area_name"),
        }
        for r in zip_rows
    }
    assign_map = {
        a["subject_id"]: a.get("zcta")
        for a in assignments
        if a.get("subject_id") and a.get("zcta")
    }

    built_at = utc_now()
    source_health = {
        "source_id": safmr_source.SOURCE_ID,
        "artifact_id": cfg.get("artifact_id") or safmr_source.ARTIFACT_ID,
        "fiscal_year": cfg.get("fiscal_year"),
        "period_start": cfg.get("period_start"),
        "period_end": cfg.get("period_end"),
        "effective_date": cfg.get("effective_date"),
        "revision": cfg.get("revision"),
        "gross_or_net": "gross",
        "measure_basis": "regulatory_market_benchmark",
        "statistic": "40th_percentile_methodology",
        "display_label": cfg.get("display_label"),
        "not_a_label": cfg.get("not_a_label"),
        "raw_snapshot": {
            "path": str(path.relative_to(project_root())) if path.exists() else None,
            "sha256": sha256_file(path) if path.exists() else None,
            "source_url": safmr_source.source_cfg().get("bulk_url"),
            "landing_page": safmr_source.source_cfg().get("landing_page"),
            "retrieved_at": receipt.get("retrieved_at"),
        },
        "zcta_snapshot": {
            "path": str(zcta_source.raw_path().relative_to(project_root()))
            if zcta_source.raw_path().exists()
            else None,
            "sha256": sha256_file(zcta_source.raw_path())
            if zcta_source.raw_path().exists()
            else None,
            "source_url": zcta_source.source_cfg().get("geojson_url", "").split("?")[0],
            "landing_page": zcta_source.source_cfg().get("landing_page"),
            "retrieved_at": zcta_receipt.get("retrieved_at"),
            "vintage": "2020",
        },
        "zip_count": len(zip_rows),
        "observation_count": len(market_obs),
        "zcta_features": len(choropleth.get("features") or []),
        "zcta_matched": (choropleth.get("meta") or {}).get("matched_zctas"),
        "zcta_missing_safmr": (choropleth.get("meta") or {}).get("missing_zctas"),
        "developments_assigned": sum(1 for a in assignments if a.get("zcta")),
        "developments_unassigned": sum(1 for a in assignments if not a.get("zcta")),
        "built_at": built_at.isoformat(),
        "api_token_required": False,
        "browser_api": False,
    }

    # Spot-check Fulton ZIP 10011 2BR (measured from official file)
    fulton_check = None
    if "10011" in by_zip:
        brs = by_zip["10011"]["bedrooms"]
        fulton_check = {
            "zip": "10011",
            "safmr_2br": brs.get("2"),
            "safmr_0br": brs.get("0"),
            "safmr_1br": brs.get("1"),
            "safmr_3br": brs.get("3"),
            "safmr_4br": brs.get("4"),
            "fiscal_year": cfg.get("fiscal_year"),
            "period_start": cfg.get("period_start"),
            "period_end": cfg.get("period_end"),
        }

    coverage = {
        "zip_count": len(zip_rows),
        "zcta_features": len(choropleth.get("features") or []),
        "zcta_with_safmr": (choropleth.get("meta") or {}).get("matched_zctas"),
        "zcta_missing_safmr": (choropleth.get("meta") or {}).get("missing_zctas"),
        "missing_zips": sorted(
            f["properties"]["zip"]
            for f in choropleth.get("features") or []
            if (f.get("properties") or {}).get("safmr_missing")
        ),
        "developments_assigned": source_health["developments_assigned"],
        "fulton_zip_check": fulton_check,
    }

    result = {
        "zip_rows": zip_rows,
        "by_zip": by_zip,
        "market_areas": market_areas,
        "market_rent_observations": market_obs,
        "zcta_choropleth": choropleth,
        "geography_assignments": assignments,
        "development_zcta": assign_map,
        "source_health": source_health,
        "coverage": coverage,
        "source_artifacts": [receipt, zcta_receipt],
        "fiscal_year": cfg.get("fiscal_year"),
        "period_start": cfg.get("period_start"),
        "period_end": cfg.get("period_end"),
        "display_label": cfg.get("display_label"),
        "not_a_label": cfg.get("not_a_label"),
        "bedroom_keys": list(BEDROOM_KEYS),
    }

    if write:
        payload_compact = {
            "fiscal_year": result["fiscal_year"],
            "period_start": result["period_start"],
            "period_end": result["period_end"],
            "effective_date": cfg.get("effective_date"),
            "display_label": result["display_label"],
            "not_a_label": result["not_a_label"],
            "measure_basis": "regulatory_market_benchmark",
            "gross_or_net": "gross",
            "statistic": "40th_percentile_methodology",
            "source_id": safmr_source.SOURCE_ID,
            "source_artifact_id": cfg.get("artifact_id") or safmr_source.ARTIFACT_ID,
            "source_url": safmr_source.source_cfg().get("landing_page"),
            "by_zip": by_zip,
            "development_zcta": assign_map,
            "geography_assignments": assignments,
            "bedroom_keys": list(BEDROOM_KEYS),
            "coverage": coverage,
            "source_health": source_health,
        }
        # Full observations are large — store separately; public mirror gets compact + layer
        for out_dir in _public_mirrors():
            out_dir.mkdir(parents=True, exist_ok=True)
            write_json(out_dir / "safmr_by_zip.json", payload_compact)
            write_json(out_dir / "source_health.json", source_health)
            write_json(out_dir / "coverage.json", coverage)
            write_json(out_dir / "geography_assignments.json", {"assignments": assignments})
            # Choropleth GeoJSON (compact separators for web)
            geo_path = out_dir / "zcta_safmr.geojson"
            text = json.dumps(
                {
                    "type": "FeatureCollection",
                    "features": choropleth["features"],
                },
                separators=(",", ":"),
                ensure_ascii=False,
            )
            geo_path.write_text(text + "\n", encoding="utf-8")
            # Layer meta alongside
            write_json(out_dir / "zcta_safmr_meta.json", choropleth.get("meta") or {})
        # Processed-only full market observations (not required in demo HTML)
        write_json(_processed_root() / "market_rent_observations.json", market_obs)
        write_json(_processed_root() / "market_areas.json", market_areas)
        write_json(_processed_root() / "zip_rows.json", zip_rows)

        # Also place choropleth under geometry public path for map loading parity
        for geo_dir in (
            project_root() / "web" / "public" / "data" / "geometry",
            project_root() / "data" / "processed" / "geometry",
            project_root() / "dist" / "app" / "data" / "geometry",
            project_root() / "dist" / "data" / "geometry",
        ):
            geo_dir.mkdir(parents=True, exist_ok=True)
            dest = geo_dir / "zcta_safmr.geojson"
            dest.write_text(
                json.dumps(
                    {"type": "FeatureCollection", "features": choropleth["features"]},
                    separators=(",", ":"),
                    ensure_ascii=False,
                )
                + "\n",
                encoding="utf-8",
            )

    return result


def load_normalized() -> dict[str, Any] | None:
    """Load compact processed SAFMR package if present."""
    path = _processed_root() / "safmr_by_zip.json"
    if not path.exists():
        path = project_root() / "web" / "public" / "data" / "hud_safmr" / "safmr_by_zip.json"
    if not path.exists():
        return None
    with path.open(encoding="utf-8") as fh:
        compact = json.load(fh)
    obs_path = _processed_root() / "market_rent_observations.json"
    market_obs = []
    if obs_path.exists():
        with obs_path.open(encoding="utf-8") as fh:
            market_obs = json.load(fh)
    areas_path = _processed_root() / "market_areas.json"
    market_areas = []
    if areas_path.exists():
        with areas_path.open(encoding="utf-8") as fh:
            market_areas = json.load(fh)
    assign_path = _processed_root() / "geography_assignments.json"
    assignments = []
    if assign_path.exists():
        with assign_path.open(encoding="utf-8") as fh:
            assignments = (json.load(fh) or {}).get("assignments") or []
    geo_path = project_root() / "data" / "processed" / "geometry" / "zcta_safmr.geojson"
    if not geo_path.exists():
        geo_path = project_root() / "web" / "public" / "data" / "geometry" / "zcta_safmr.geojson"
    choropleth = None
    if geo_path.exists():
        with geo_path.open(encoding="utf-8") as fh:
            choropleth = json.load(fh)
    return {
        "by_zip": compact.get("by_zip") or {},
        "development_zcta": compact.get("development_zcta") or {},
        "geography_assignments": assignments or compact.get("geography_assignments") or [],
        "market_rent_observations": market_obs,
        "market_areas": market_areas,
        "zcta_choropleth": choropleth,
        "source_health": compact.get("source_health"),
        "coverage": compact.get("coverage"),
        "fiscal_year": compact.get("fiscal_year"),
        "period_start": compact.get("period_start"),
        "period_end": compact.get("period_end"),
        "display_label": compact.get("display_label"),
        "not_a_label": compact.get("not_a_label"),
        "bedroom_keys": compact.get("bedroom_keys") or list(BEDROOM_KEYS),
        "source_artifact_id": compact.get("source_artifact_id"),
        "source_url": compact.get("source_url"),
        "source_id": compact.get("source_id") or safmr_source.SOURCE_ID,
    }


# Silence unused import warning for mapping if tree path unused
_ = mapping
_ = Counter
